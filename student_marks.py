import openpyxl

# Function to calculate average grade
def calculate_per(marks):      # list of marks passed as argument
    total = sum(marks)
    result = (total / 400) * 100
    return result

# Define the Excel file name
excel_file = "Marklist.xlsx"

try:
    # Try to load an existing workbook
    workbook = openpyxl.load_workbook(excel_file)
except FileNotFoundError:
    # If the file doesn't exist, create a new workbook
    workbook = openpyxl.Workbook()
    workbook.save(excel_file)  # Save the new workbook to the file

# Main loop
while True:
    print("\nOPTIONS")
    print("1. Add Class and Students")
    print("2. Generate Reports")
    print("3. Save and Exit")
    
    choice = input("\nWhat do you wish to do? (Please enter appropriate option): ")
    
    if choice == "1":
        num_classes = int(input("Enter the number of classes: "))
        
        for _ in range(num_classes):
            class_name = input("Enter class: ")
            num_students = int(input(f"Enter number of students in {class_name}: "))
            
            if class_name in workbook.sheetnames:
                class_sheet = workbook[class_name]  # Open the existing sheet
            else:
                class_sheet = workbook.create_sheet(title=class_name)  # Create a new sheet
                class_sheet.append(["Student ID", "Student Name", "Computer", "Math", "Science", "English", "Percentage"])
            
            for i in range(num_students):
                print("Enter details of Student {}-".format(i+1))
                student_id = int(input("\nStudent ID: "))
                student_name = input("Name: ")
                marks = []
                print("\nEnter marks out of 100.")
                for subject in ["Computer", "Math", "Science", "English"]:
                    marks.append(float(input(f"{subject}: ")))
                percentage = calculate_per(marks)
                class_sheet.append([student_id, student_name] + marks + [percentage])
                print()
            print("---------------------------------------------------------------")

    elif choice == "2":
        while True:
            print('''\nREPORT MENU:"
    1. Display Student Report
    2. Display Class Statistics
    3. Display all Student Reports in a Class
    4. Back to Main Menu ''')
            
            ch = input("\nEnter your choice: ")
            
            if ch == "1":
                student_name = input("\nEnter student name to generate academic report: ")
                class_name = input("Enter class: ")
                
                if class_name in workbook.sheetnames:      # list of sheets in workbook
                    class_sheet = workbook[class_name]
                    student_found = False
                    
                    for row in class_sheet.iter_rows(min_row=2, values_only=True):      # 1st row headers, start from 2nd
                        id, name, *marks, per = row         # unpacking, *grades stores all remaining values
                        if name.title() == student_name.title():
                            print()
                            print(f"Student ID: {id}")
                            print(f"Name: {name}")
                            print(f"Marks: {marks}")
                            print(f"Percentage: {per:.2f}")
                            student_found = True
                            break
                    
                    if not student_found:
                        print("Student not found.")
                else:
                    print("Class not found.")
                print("---------------------------------------------------------------")
                
            elif ch == "2":
                class_name = input("\nEnter class: ")
                
                if class_name in workbook.sheetnames:
                    class_sheet = workbook[class_name]
                    percentage_list = []
                    total_students = 0
                    total_per = 0
                    highest_percentage = 0
                    highest_percentage_students = []
                    
                    for row in class_sheet.iter_rows(min_row=2, values_only=True):
                        student_id, student_name, *marks, percentage = row
                        total_students += 1
                        total_per += percentage
                        percentage_list.append(percentage)
                        
                        if percentage > highest_percentage:
                            highest_percentage = percentage
                            highest_percentage_students = [(student_id, student_name)]
                        elif percentage == highest_percentage:
                            highest_percentage_students.append((student_id, student_name))
                    
                    class_average = total_per / total_students
                    print(f"\nClass: {class_name}")
                    print(f"Highest Percentage: {highest_percentage:.2f} scored by:")
                    for student_id, student_name in highest_percentage_students:
                        print(f"-> {student_name} ({student_id})")
                    print(f"Class Average: {class_average:.2f}")
                
                else:
                    print("Class not found.")
                
                print("-" * 100)


            elif ch == "3":
                class_name = input("\nEnter class: ")
                
                if class_name in workbook.sheetnames:
                    class_sheet = workbook[class_name]
                    
                    for row in class_sheet.iter_rows(min_row=2, values_only=True):
                        print("Student Details for Class:", class_name)
                        print("-" * 100)
                        print("{:<10} {:<20} {:<10} {:<10} {:<10} {:<10} {:<12}".format(
                            "ID", "Name", "Computer", "Math", "Science", "English", "Percentage"))
                        print("=" * 100)
                        
                        for row in class_sheet.iter_rows(min_row=2, values_only=True):
                            student_id, student_name, computer, math, science, english, percentage = row
                            print("{:<10} {:<20} {:<10} {:<10} {:<10} {:<10} {:<12}".format(
                                student_id, student_name, computer, math, science, english, percentage))
        
                else:
                    print("Class not found.")
                print("---------------------------------------------------------------")

            elif ch == "4":
                print("Back to main menu...")
                print("---------------------------------------------------------------")

                break

            else:
                print("Invalid choice. Please enter a valid option.")
    
    elif choice == "3":
        workbook.save("Marklist.xlsx")
        print("\nWorkbook saved. Exiting.")
        break
    
    else:
        print("\nInvalid choice. Please enter a valid option.")
